"""
Quarter-on-Quarter DST(CFARRI) Comparison Report Generator
"""

import pandas as pd
from pathlib import Path
from itertools import product
import ast
import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


# ============================================================
# CONFIGURATION
# ============================================================

OLD_QUARTER = "Q2_25"
NEW_QUARTER = "Q3_25"

business_keys_list = [
    "CF_FINAL_ATTRIBUTE",
    "CF_ACTUAL_SOURCE",
    "CF_SET_CODE",
]

transformation_prefix = "CF_SRC_TGT_TRANSFORMATION"
MIN_PREFIX_LEN = 3


# ============================================================
# FILE READING
# ============================================================

def read_lineage_file(file_path: str) -> pd.DataFrame:

    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(file_path)

    if file_path.suffix.lower() == ".csv":
        return pd.read_csv(file_path)

    elif file_path.suffix.lower() in (".xlsx", ".xls"):
        return pd.read_excel(file_path)

    else:
        raise ValueError("Unsupported file format")


# ============================================================
# NORMALIZATION
# ============================================================

def normalize_transformation_dtypes(df, prefix):

    for c in df.columns:
        if c.startswith(prefix):
            df[c] = df[c].astype("string")

    return df


# ============================================================
# REMOVE FULLY MATCHING RECORDS
# ============================================================

def remove_fully_matching_records(
        old_df,
        new_df,
        business_keys
):

    trans_cols = [
        c for c in old_df.columns
        if c.startswith(transformation_prefix)
    ]

    compare_cols = business_keys + trans_cols

    matched_df = old_df.merge(
        new_df,
        on=compare_cols,
        how="inner"
    )

    old_remaining = (
        old_df.merge(
            matched_df[compare_cols],
            on=compare_cols,
            how="left",
            indicator=True
        )
        .query('_merge=="left_only"')
        .drop(columns="_merge")
    )

    new_remaining = (
        new_df.merge(
            matched_df[compare_cols],
            on=compare_cols,
            how="left",
            indicator=True
        )
        .query('_merge=="left_only"')
        .drop(columns="_merge")
    )

    return old_remaining, new_remaining, matched_df


# ============================================================
# CORE COMPARISON
# ============================================================

def compare_remaining_lineage_cartesian(
        old_df,
        new_df,
        business_keys
):

    trans_cols = sorted(
        c for c in old_df.columns
        if c.startswith(transformation_prefix)
    )

    section1_rows = []

    for key_vals, old_grp in old_df.groupby(
            business_keys
    ):

        key_dict = dict(
            zip(
                business_keys,
                key_vals
            )
        )

        new_grp = new_df[
            (
                new_df[business_keys]
                ==
                pd.Series(key_dict)
            ).all(axis=1)
        ]

        if new_grp.empty:
            continue

        for old_row, new_row in product(
                old_grp.to_dict("records"),
                new_grp.to_dict("records")
        ):

            row_out = key_dict.copy()

            changed_nums = []

            for col in trans_cols:

                num = col.split(
                    transformation_prefix + "_"
                )[-1]

                old_val = (
                    ""
                    if pd.isna(old_row[col])
                    else str(old_row[col])
                )

                new_val = (
                    ""
                    if pd.isna(new_row[col])
                    else str(new_row[col])
                )

                row_out[
                    f"{transformation_prefix}_{num} (Old)"
                ] = old_val

                row_out[
                    f"{transformation_prefix}_{num} (New)"
                ] = new_val

                if old_val != new_val:
                    changed_nums.append(num)

            if changed_nums:

                row_out[
                    "Transformation Number"
                ] = changed_nums

                section1_rows.append(
                    row_out
                )

    section1_df = pd.DataFrame(
        section1_rows
    )

    section2_df = new_df[
        ~new_df[business_keys]
        .apply(tuple, 1)
        .isin(
            old_df[business_keys]
            .apply(tuple, 1)
        )
    ]

    section3_df = old_df[
        ~old_df[business_keys]
        .apply(tuple, 1)
        .isin(
            new_df[business_keys]
            .apply(tuple, 1)
        )
    ]

    return (
        section1_df,
        section2_df,
        section3_df
    )


# ============================================================
# CANONICAL COLUMN NORMALIZATION
# ============================================================

def normalize_newly_added(df):

    df = df.copy()

    for c in list(df.columns):

        if (
                c.startswith(
                    transformation_prefix
                )
                and "(" not in c
        ):

            df[f"{c} (New)"] = df[c]
            df[f"{c} (Old)"] = ""

            df.drop(
                columns=c,
                inplace=True
            )

    return df


def normalize_removed(df):

    df = df.copy()

    for c in list(df.columns):

        if (
                c.startswith(
                    transformation_prefix
                )
                and "(" not in c
        ):

            df[f"{c} (Old)"] = df[c]
            df[f"{c} (New)"] = ""

            df.drop(
                columns=c,
                inplace=True
            )

    return df


def align_columns(
        df,
        all_columns
):

    for c in all_columns:
        if c not in df.columns:
            df[c] = ""

    return df[
        list(all_columns)
    ]


# ============================================================
# INTERLEAVED COLUMN ORDER
# ============================================================

def build_interleaved_column_order(
        final_df,
        new_file_columns_order
):

    ordered_cols = []
    inserted_diff = False

    for col in new_file_columns_order:

        if col in business_keys_list:
            ordered_cols.append(col)
            continue

        if not inserted_diff:

            ordered_cols.append(
                "Difference"
            )

            ordered_cols.append(
                "Transformation Number"
            )

            inserted_diff = True

        if col.startswith(
                transformation_prefix + "_"
        ):

            num = col.split(
                transformation_prefix + "_"
            )[-1]

            ordered_cols.append(
                f"{transformation_prefix}_{num} (Old)"
            )

            ordered_cols.append(
                f"{transformation_prefix}_{num} (New)"
            )

        else:
            ordered_cols.append(col)

    return [
        c
        for c in ordered_cols
        if c in final_df.columns
    ]


# ============================================================
# EXCEL COLORING
# ============================================================

NEW_LOGIC_FILL = PatternFill(
    "solid",
    fgColor="FFE4B084"
)

OLD_LOGIC_FILL = PatternFill(
    "solid",
    fgColor="FFFFC000"
)

RED_BOLD_FONT = InlineFont(
    color="FF0000",
    b=True
)

NORMAL_FONT = InlineFont(
    color="000000"
)


def common_prefix_length(a, b):

    i = 0

    while (
            i < min(len(a), len(b))
            and a[i] == b[i]
    ):
        i += 1

    return i


def apply_complete_change_color(
        ws,
        row,
        new_col,
        old_col
):

    ws.cell(
        row=row,
        column=new_col
    ).fill = NEW_LOGIC_FILL

    ws.cell(
        row=row,
        column=old_col
    ).fill = OLD_LOGIC_FILL


def apply_partial_change(
        ws,
        row,
        new_col,
        old_col,
        old_text,
        new_text,
        prefix_len
):

    def build(text):

        parts = []

        if text[:prefix_len]:
            parts.append(
                TextBlock(
                    NORMAL_FONT,
                    text[:prefix_len]
                )
            )

        if text[prefix_len:]:
            parts.append(
                TextBlock(
                    RED_BOLD_FONT,
                    text[prefix_len:]
                )
            )

        return (
            CellRichText(parts)
            if parts
            else text
        )

    ws.cell(
        row=row,
        column=new_col
    ).value = build(new_text)

    ws.cell(
        row=row,
        column=old_col
    ).value = build(old_text)


def apply_excel_coloring(
        excel_path
):

    wb = load_workbook(
        excel_path
    )

    ws = wb[
        "Lineage_Comparison"
    ]

    headers = {
        c.value: i + 1
        for i, c in enumerate(
            ws[1]
        )
    }

    for r in range(
            2,
            ws.max_row + 1
    ):

        if (
                ws.cell(
                    r,
                    headers[
                        "Difference"
                    ]
                ).value
                !=
                "Transformation_Logic_Change"
        ):
            continue

        nums = ast.literal_eval(
            str(
                ws.cell(
                    r,
                    headers[
                        "Transformation Number"
                    ]
                ).value
            )
        )

        for n in nums:

            new_col = headers.get(
                f"{transformation_prefix}_{n} (New)"
            )

            old_col = headers.get(
                f"{transformation_prefix}_{n} (Old)"
            )

            if (
                    not new_col
                    or
                    not old_col
            ):
                continue

            new_val = str(
                ws.cell(
                    r,
                    new_col
                ).value or ""
            )

            old_val = str(
                ws.cell(
                    r,
                    old_col
                ).value or ""
            )

            prefix = common_prefix_length(
                old_val,
                new_val
            )

            if prefix < MIN_PREFIX_LEN:

                apply_complete_change_color(
                    ws,
                    r,
                    new_col,
                    old_col
                )

            else:

                apply_partial_change(
                    ws,
                    r,
                    new_col,
                    old_col,
                    old_val,
                    new_val,
                    prefix
                )

    wb.save(excel_path)


def align_transformation_columns(
        old_df,
        new_df,
        prefix
):

    old_trans_cols = [
        c for c in old_df.columns
        if c.startswith(prefix)
    ]

    new_trans_cols = [
        c for c in new_df.columns
        if c.startswith(prefix)
    ]

    all_trans_cols = sorted(
        set(old_trans_cols)
        |
        set(new_trans_cols)
    )

    for col in all_trans_cols:

        if col not in old_df.columns:
            old_df[col] = ""

        if col not in new_df.columns:
            new_df[col] = ""

    return old_df, new_df


def reorder_df_by_columns_hierarchy(
        unordered_df
):

    df = unordered_df

    original_cols = (
        df.columns.tolist()
    )

    new_cols = []
    used = set()

    prefix = "CF_SRC_TGT_TRANSFORMATION_"

    for col in original_cols:

        if col in used:
            continue

        new_cols.append(col)
        used.add(col)

        if col.startswith(prefix):

            suffix = col[
                     len(prefix):
                     ]

            if suffix.isdigit():

                children_prefix = (
                        col + "_"
                )

                children = [
                    c
                    for c in original_cols
                    if (
                            c.startswith(
                                children_prefix
                            )
                            and c not in used
                    )
                ]

                children.sort()

                new_cols.extend(
                    children
                )

                used.update(
                    children
                )

    for col in original_cols:

        if col not in used:
            new_cols.append(
                col
            )

    ordered_df = df[new_cols]

    return ordered_df


# ============================================================
# FOLDER CONFIGURATION
# ============================================================

OLD_FOLDER = "Old"
NEW_FOLDER = "New"
REPORTS_FOLDER = "Reports"
DST_SUBFOLDER = "DST_Comparison"


# ============================================================
# MAIN
# ============================================================

def main():

    print(
        "----- DST Comparison Utility -----"
    )

    dst_comparison_folder = (
        os.path.join(
            REPORTS_FOLDER,
            DST_SUBFOLDER
        )
    )

    print(
        "DST folder path:",
        dst_comparison_folder
    )

    os.makedirs(
        dst_comparison_folder,
        exist_ok=True
    )

    try:

        old_files = set(
            os.listdir(
                OLD_FOLDER
            )
        )

        new_files = set(
            os.listdir(
                NEW_FOLDER
            )
        )

    except FileNotFoundError:

        print(
            f"Error: "
            f"Directory not found. "
            f"Please ensure "
            f"{OLD_FOLDER} "
            f"and "
            f"{NEW_FOLDER} exist."
        )

        return

    diff_report_path = os.path.join(
        REPORTS_FOLDER,
        "File_Difference_Report.xlsx"
    )

    with pd.ExcelWriter(
            diff_report_path
    ) as writer:

        pd.DataFrame(
            list(
                old_files - new_files
            ),
            columns=[
                "Removed Files"
            ]
        ).to_excel(
            writer,
            sheet_name="Removed Files",
            index=False
        )

        pd.DataFrame(
            list(
                new_files - old_files
            ),
            columns=[
                "Added Files"
            ]
        ).to_excel(
            writer,
            sheet_name="Added Files",
            index=False
        )

    common_files = sorted(
        list(
            old_files
            &
            new_files
        )
    )

    if not common_files:

        print(
            "No common files found."
        )

        return

    print(
        f"Found "
        f"{len(common_files)} "
        f"common files"
    )

    for common_file in common_files:

        old_file_path = os.path.join(
            OLD_FOLDER,
            common_file
        )

        new_file_path = os.path.join(
            NEW_FOLDER,
            common_file
        )

        report_filename = (
            f"{os.path.splitext(common_file)[0]}"
            f"_report.xlsx"
        )

        output_report_path = (
            os.path.join(
                dst_comparison_folder,
                report_filename
            )
        )

        old_df = (
            read_lineage_file(
                old_file_path
            ).fillna("")
        )

        new_df = (
            read_lineage_file(
                new_file_path
            ).fillna("")
        )

        old_df = normalize_transformation_dtypes(
            old_df,
            transformation_prefix
        )

        new_df = normalize_transformation_dtypes(
            new_df,
            transformation_prefix
        )

        old_df, new_df = (
            align_transformation_columns(
                old_df,
                new_df,
                transformation_prefix
            )
        )

        old_df = reorder_df_by_columns_hierarchy(
            old_df
        )

        new_df = reorder_df_by_columns_hierarchy(
            new_df
        )

        new_file_columns_order = (
            list(
                new_df.columns
            )
        )

        old_remain, new_remain, _ = (
            remove_fully_matching_records(
                old_df,
                new_df,
                business_keys_list
            )
        )

        change_df, add_df, rem_df = (
            compare_remaining_lineage_cartesian(
                old_remain,
                new_remain,
                business_keys_list
            )
        )

        change_df[
            "Difference"
        ] = (
            "Transformation_Logic_Change"
        )

        add_df[
            "Difference"
        ] = (
            f"Newly Added "
            f"in {NEW_QUARTER}"
        )

        rem_df[
            "Difference"
        ] = (
            f"Removed from "
            f"{NEW_QUARTER}"
        )

        add_df = normalize_newly_added(
            add_df
        )

        rem_df = normalize_removed(
            rem_df
        )

        all_cols = (
                set(change_df.columns)
                |
                set(add_df.columns)
                |
                set(rem_df.columns)
        )

        final_df = pd.concat(
            [
                align_columns(
                    change_df,
                    all_cols
                ),

                align_columns(
                    add_df,
                    all_cols
                ),

                align_columns(
                    rem_df,
                    all_cols
                )
            ],
            ignore_index=True
        )

        final_df = final_df[
            build_interleaved_column_order(
                final_df,
                new_file_columns_order
            )
        ]

        with pd.ExcelWriter(
                output_report_path,
                engine="openpyxl"
        ) as writer:

            final_df.to_excel(
                writer,
                sheet_name="Lineage_Comparison",
                index=False
            )

        apply_excel_coloring(
            output_report_path
        )


if __name__ == "__main__":
    main()